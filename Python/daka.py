import openpyxl
import os
import time
path = os.getcwd()
current_list = os.listdir(path)
gmtime = time.gmtime()
suffix = "%d%d%d"%(gmtime.tm_year,gmtime.tm_mon,gmtime.tm_mday)
file_name = ""
white_list = ["王渤轩"]
for i in current_list:
    if i.__contains__(suffix):
        file_name = i
if file_name == "":
    print(f"No file found using {suffix}")
    exit(0)

file = openpyxl.open(file_name)
sheet = file.active
row = 0
_dict = {}
while True:
    row += 1
    cell = sheet.cell(row, 3)
    if cell.value is None:
        print("Traverse Done.")
        break
    if cell.value.__contains__("190") and sheet.cell(row,1).value not in white_list:
        _dict[sheet.cell(row,1).value] = cell.value

_list = sorted(_dict.items(),key = lambda x:(x[1],x[0]))
print(len(_list))

row = 0
for i in _list:
    row += 1
    sheet.cell(row, 10).value = i[0]
    sheet.cell(row, 11).value = i[1]
sheet.column_dimensions['K'].width = 24
file.save(file_name)
file.close()