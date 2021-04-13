# Author: c01dkit
# Create time: 2020-11-10
# last updated: 2020-11-10
import openpyxl

# The following names must be assigned.
src_file_name = "成绩单发快递.xlsx" # 必须修改，指定数据源文件的全名： xxxx.xslx
src_c2 = [5,6,4] # 必须修改，指定数据源用于填充的数据段所在的列(1,2,3,...)
src_B = 3 # 必须修改，指定数据源用于和目标源相匹配的列
src_sheet_name = "Sheet2" # 看情况要不要修改，选择表单名称
des_file_name = "2019级学生名单_整理20210412_bak.xlsx"  # 必须修改，指定目标源文件的全名： xxxx.xslx
des_c1 = [6,7,8] # 必须修改，指定目标源需要被填充的列(1,2,3,...)
des_A = 2 # 必须修改，指定目标源用于和数据源相匹配的列
des_sheet_name = "Sheet2" # 看情况要不要修改，选择表单名称
version = "_v2" # 版本号，可以修改
save_file_name = des_file_name[:-5] + version + ".xlsx"
des_row = 1 # 从第des_row+1处开始遍历每行, des_row=1时默认首行为标题不匹配
def fill_des_c1_on_A_with_src_c2_on_B(des, c1, A, src, c2, B, des_row):
    """This method will fill des'c1 column with src'c2 column from des_row+1 row
    if the values of des'A and src'B are the same.
    """
    if len(c1) != len(c2):
        print("指定需要同步的数据列数不相等！")
        exit(0)
    try:
        src_file = openpyxl.open(src)
    except FileNotFoundError:
        print(f"当前脚本所在目录没有找到 \"{src}\"")
        return
    try:
        src_sheet = src_file[src_sheet_name]
    except KeyError:
        print(f"\"{src}\" 不具有 \"{src_sheet_name}\" 这一表单名")
        return
    try:
        des_file = openpyxl.open(des)
    except FileNotFoundError:
        print(f"当前脚本所在目录没有找到 \"{src}\"")
        return
    try:
        des_sheet = des_file[des_sheet_name]
    except KeyError:
        print(f"\"{des}\" 不具有 \"{src_sheet_name}\" 这一表单名")
        return
    for i in range(len(c1)):
        # 循环遍历des的A列，直到遇到一个空值退出循环。（A是用于连接src的B列的键）
        found = 0
        not_found = 0
        des_row_start = des_row
        while True:
            des_row_start += 1
            des_cell_A = des_sheet.cell(row=des_row_start, column=A)
            if des_cell_A.value is None:
                print(des_file_name+f" 的第 {A} 列已经遍历完毕，合并完成。共合并 {found} 项数据，缺少 {not_found} 项数据")
                break
            # 如果对A列部分字段有需要跳过，可以在这里声明
            # if str(des_cell_A.value):
            #     continue
            des_cell_c1 = des_sheet.cell(row=des_row_start, column=c1[i])
            # 如果此项不为空，跳过。
            if des_cell_c1.value is not None:
                continue
            src_row = 0 # 全列搜索
            # 遍历数据源，寻找匹配项
            flag = False
            while True:
                src_row += 1
                src_cell_B = src_sheet.cell(row=src_row, column=B)
                if src_cell_B.value is None:
                    if not flag:
                        print(f"{src_file_name} 缺少 {des_cell_A.value} 需要的第 {src_c2} 列内容")
                    break
                if str(src_cell_B.value) == str(des_cell_A.value):
                    src_cell_c2 = src_sheet.cell(row=src_row, column=c2[i])
                    if src_cell_c2.value is None:
                        not_found += 1
                        print(f"{src_file_name} 具有 {des_cell_A.value}，但其第 {src_c2} 列内容为空，请确认")
                        break
                    des_cell_c1.value = src_cell_c2.value
                    found += 1
                    flag = True
    try:
        des_file.save(save_file_name)
    except PermissionError:
        print(f"\n保存失败。请不要在脚本运行时打开\"{save_file_name}\"")
    des_file.close()
    src_file.close()

if __name__ == '__main__':
    fill_des_c1_on_A_with_src_c2_on_B(des_file_name, des_c1, des_A,
                                      src_file_name, src_c2, src_B, des_row)