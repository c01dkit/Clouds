import os
import openpyxl
import re
import numpy as np

from openpyxl.styles import Alignment, Font

"""
本jio本用于信安1801班2020年班级评议的得分统计
时间线:
    - 2020年8月9日
        - 增加了方差计算
        - 修改了运行须知
    - 2020年8月8日
        - 提交初稿
    - 2021年8月19日
        - 更新注释
运行须知:
    1. jio本需要放置于数据源所在的目录下
    2. 数据源必须以"2021班级评议-xxx.xlsx"命名,否则不参与计算
    3. 计算得分时,自己的四项得分不参与计算;他人的班级活动评价不参与计算
    4. 应填但未填的部分在处理的时候按满分计算, 填写但是不在范围内的也按满分计算
    5. 单项评价满分分别为 20 20 20 10 自评按 30 分计算
    6. 输出结果按总得分降序排列
    7. 输出结果单元格没有使用Excel自带的sum函数.如果有变更请重新运行jio本
    8. 重复运行jio本将原地新建或覆盖
    9. 每处理一份评价表,输出学生对他人评价各项均分, 总分均分, 以及方差. 方差越大说明该份评价越有区分度
    10. 程序结束时输出没有提交或者把自己名字打错的成员名单
    11. 运行程序时不要打开最终结果xlsx, 否则提示Permisson Denied
输出结果:
    2021信安1801班级评议统计表.xlsx
"""


class Macro:
    EXECUTED_WORDS = "处理完成,结果保存在 "
    INITIATE_DONE = "总评表初始化完成"
    RESULT_ATTR = ["学号", "姓名", "学习勤奋刻苦", "积极奉献班级", "团结帮助同学",
                   "其他总分", "总分1", "班级活动自评", "总分2", "排名"]
    RESULT_AVAILABLE_COLUMNS = "ABCDEFGHIJ"
    RESULT_TITLE_NAME = "2021信安1801班级评议统计表"
    RESULT_FILE_NAME = "2021信安1801班级评议统计表.xlsx"
    SOURCE_AVG = [" 对他人评价均分情况-> ", "学习勤奋刻苦:", " 积极奉献班级:", " 团结帮助同学:", " 其他:", " 总分:"]
    SOURCE_FILE_STARTS_WITH = "2021班级评议"
    SOURCE_HANDLE = "提交的评价表处理完毕 "
    SOURCE_NOT_FOUND = "没有找到数据源,请检查文件路径"
    SOURCE_ROW_STARTS_AT = 12   # 起始有效行数序号
    SOURCE_ROW_END_AT = 40      # 需要修改为最后有效行数序号-1
    SOURCE_VAR = ["学习勤奋刻苦评分方差:", " 积极奉献班级评分方差:", " 团结帮助同学方差:", " 其他评分方差:", " 各项方差和:"]


def input_valid(value, limit):
    if type(value) == int:
        if value < 0 or value > limit:
            value = limit
    else:
        value = limit
    return value


class Student:
    all_stu = []

    def __init__(self, sno, name, struggle=0, dedication=0, unity=0, misc=0, participation=0):
        self.total = 0
        self.sno = sno
        self.name = name
        self.struggle = struggle
        self.dedication = dedication
        self.unity = unity
        self.misc = misc
        self.participation = participation
        self.all_stu.append(name)

    def add_score(self, struggle=0, dedication=0, unity=0, misc=0, participation=0):
        struggle = input_valid(struggle, 20)
        self.struggle += struggle
        dedication = input_valid(dedication, 20)
        self.dedication += dedication
        unity = input_valid(unity, 20)
        self.unity += unity
        misc = input_valid(misc, 10)
        self.misc += misc
        participation = input_valid(participation, 30)
        self.participation += participation

    def calculate_total_score(self):
        self.total = self.struggle + self.dedication + self.unity + self.participation + self.misc


def init_result():
    __file_name = Macro.RESULT_FILE_NAME
    __result = openpyxl.Workbook()
    __result.save(__file_name)
    __result.close()
    __result = openpyxl.load_workbook(__file_name)
    __attr = Macro.RESULT_ATTR
    __sheet = __result.active
    __co = list(Macro.RESULT_AVAILABLE_COLUMNS)
    for __i in __co:
        __sheet.column_dimensions[__i].width = 15
    __sheet.row_dimensions[1].height = 30
    __sheet.row_dimensions[2].height = 20
    __title_cell = __sheet.cell(1, 1, Macro.RESULT_TITLE_NAME)
    __title_cell.font = Font(bold=True, size=16)
    __title_cell.alignment = Alignment(horizontal="center", vertical="center")
    __sheet.merge_cells("A1:J1")
    for __i in range(1, len(Macro.RESULT_AVAILABLE_COLUMNS) + 1):
        __attr_cell = __sheet.cell(2, __i, __attr[__i - 1])
        __attr_cell.alignment = Alignment(horizontal="center", vertical="center")
    __result.save(__file_name)
    __result.close()
    print(Macro.INITIATE_DONE)
    return __file_name


def get_file_list():
    __files_name = []
    __path = os.getcwd()
    __all_files = os.listdir(__path)
    for __name in __all_files:
        __name_split = re.split('[-.]', __name)
        if len(__name_split) == 3 \
                and __name_split[0] == Macro.SOURCE_FILE_STARTS_WITH \
                and __name_split[2] == "xlsx":
            __files_name.append(__name)

    return __files_name


def load_data(__students, __file_name):  # __students是全体成员,不是仅提交评价表的部分
    __row_start = Macro.SOURCE_ROW_STARTS_AT
    __file = openpyxl.load_workbook(__file_name)
    __sheet = __file.active
    __eval_name = re.split('[-.]', __file_name)[1]
    Student.all_stu.remove(__eval_name)
    for __student in __students:
        if __student.name == __eval_name:
            __student.add_score(participation=input_valid(__sheet.cell(column=8, row=__row_start).value,30))
            __row_start += 1
        else:
            __student.add_score(struggle=input_valid(__sheet.cell(column=3, row=__row_start).value,20),
                                dedication=input_valid(__sheet.cell(column=4, row=__row_start).value,20),
                                unity=input_valid(__sheet.cell(column=5, row=__row_start).value,20),
                                misc=input_valid(__sheet.cell(column=6, row=__row_start).value,10))
            __row_start += 1
    print(__eval_name + Macro.SOURCE_HANDLE, end="")
    __tmp1, __tmp2, __tmp3, __tmp4 = [], [], [], []
    for __row in range(Macro.SOURCE_ROW_STARTS_AT, Macro.SOURCE_ROW_END_AT):
        if __sheet.cell(__row, 2).value == __eval_name:
            print("自评%d/30分" % input_valid(__sheet.cell(__row, 8).value, 30), end=" ")
        else:
            __tmp1.append(input_valid(__sheet.cell(__row, 3).value,20))
            __tmp2.append(input_valid(__sheet.cell(__row, 4).value,20))
            __tmp3.append(input_valid(__sheet.cell(__row, 5).value,20))
            __tmp4.append(input_valid(__sheet.cell(__row, 6).value,10))
    __tmp1_avg = np.mean(__tmp1)
    __tmp2_avg = np.mean(__tmp2)
    __tmp3_avg = np.mean(__tmp3)
    __tmp4_avg = np.mean(__tmp4)
    __tmp1_var = np.var(__tmp1)
    __tmp2_var = np.var(__tmp2)
    __tmp3_var = np.var(__tmp3)
    __tmp4_var = np.var(__tmp4)
    print(Macro.SOURCE_AVG[0],
          Macro.SOURCE_AVG[1], "{:.2f}/20".format(__tmp1_avg), Macro.SOURCE_AVG[2], "{:.2f}/20".format(__tmp2_avg),
          Macro.SOURCE_AVG[3], "{:.2f}/20".format(__tmp3_avg), Macro.SOURCE_AVG[4], "{:.2f}/10".format(__tmp4_avg),
          Macro.SOURCE_AVG[5], "{:.2f}/70".format(__tmp1_avg + __tmp2_avg + __tmp3_avg + __tmp4_avg),
          sep="", end=" ")
    print(Macro.SOURCE_VAR[0], "{:.4f}".format(__tmp1_var),
          Macro.SOURCE_VAR[1], "{:.4f}".format(__tmp2_var),
          Macro.SOURCE_VAR[2], "{:.4f}".format(__tmp3_var),
          Macro.SOURCE_VAR[3], "{:.4f}".format(__tmp4_var),
          Macro.SOURCE_VAR[4], "{:.4f}".format(__tmp1_var + __tmp2_var + __tmp3_var + __tmp4_var),
          sep="")
    __file.close()


def sort(__students):
    for __student in __students:
        __student.calculate_total_score()
    __students.sort(key=lambda x: x.total, reverse=True)


def store_data(__students, __result_file_name):
    __result_file = openpyxl.load_workbook(__result_file_name)
    __sheet = __result_file.active
    __row_start = 3
    __rank = 0
    __step = 1
    __last_score = 9999999
    for __student in __students:
        __sheet.row_dimensions[__row_start].height = 20
        __sheet.cell(row=__row_start, column=1, value=str(__student.sno))
        __sheet.cell(row=__row_start, column=2, value=str(__student.name))
        __sheet.cell(row=__row_start, column=3, value=str(__student.struggle))
        __sheet.cell(row=__row_start, column=4, value=str(__student.dedication))
        __sheet.cell(row=__row_start, column=5, value=str(__student.unity))
        __sheet.cell(row=__row_start, column=6, value=str(__student.misc))
        __sheet.cell(row=__row_start, column=7, value=str(__student.total - __student.participation))
        __sheet.cell(row=__row_start, column=8, value=str(__student.participation))
        __sheet.cell(row=__row_start, column=9, value=str(__student.total))
        if __student.total < __last_score:
            __last_score = __student.total
            __rank = __step
        __step += 1
        __sheet.cell(row=__row_start, column=10, value=str(__rank))
        for __i in range(1, 11):
            __temp_cell = __sheet.cell(__row_start, __i)
            __temp_cell.alignment = Alignment(horizontal="center", vertical="center")
        __row_start += 1
    __result_file.save(__result_file_name)
    __result_file.close()


if __name__ == '__main__':
    result_file_name = init_result()
    files_name = get_file_list()
    if len(files_name) <= 0:
        print(Macro.SOURCE_NOT_FOUND)
        exit(0)
    students = []
    file_for_init = openpyxl.load_workbook(str(files_name[0]))
    sheet_for_init = file_for_init.active
    for i in range(Macro.SOURCE_ROW_STARTS_AT, Macro.SOURCE_ROW_END_AT):
        load_sno = sheet_for_init.cell(row=i, column=1).value
        load_name = sheet_for_init.cell(row=i, column=2).value
        students.append(Student(load_sno, load_name))
    file_for_init.close()

    for file_name in files_name:
        load_data(students, file_name)

    sort(students)
    store_data(students, result_file_name)
    print(Macro.EXECUTED_WORDS + os.path.abspath(result_file_name))
    if len(Student.all_stu) > 0:
        print("以下%d人没有提交评价表,或者评价表命名出错:" % len(Student.all_stu))
        for i in Student.all_stu:
            print(i)
    else:
        print("所有成员均提交评价表.")